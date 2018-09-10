# -*- coding: utf-8 -*-
from selenium import webdriver
import time
import random
from bs4 import BeautifulSoup as BS
from selenium.webdriver.common.keys import Keys
import pickle
from fake_useragent import UserAgent
import simplejson
from openpyxl.workbook import Workbook
import re
from interruptingcow import timeout

#Make a compiler to remove everything but numbers and decimals. Compilers take a while to make, so it's best to just do it once at the top.
non_decimal = re.compile(r'[^\d.]+')

def start_driver():
    # Start up the webdriver used to browse the web and collect data

    # Start up the fake user agent lib so that we can get a random user agent for our browser
    ua = UserAgent()

    #Add actual headers so we look legit. Also want to make sure we acceot all the right things, and that everything loads in English
    # headers = { 'Accept':'*/*',
    #     'Accept-Encoding':'gzip, deflate, sdch',
    #     'Accept-Language':'en-US,en;q=0.8',
    #     'Cache-Control':'max-age=0',
    #     'User-Agent': ua.random
    # }
    #Iterate through headers dict and add all the headers. Comment this out if you don't want to use any of it.
    # for key, value in enumerate(headers):
    #     webdriver.DesiredCapabilities.PHANTOMJS['phantomjs.page.customHeaders.{}'.format(key)] = value

    #This is my proxy. I'm using one primarily out of the Netherlands for this one. (USER/PASS here is blank because it's all private login info)
    # service_args = [
    #     '--proxy=proxy-nl.privateinternetaccess.com:1080',
    #     '--proxy-type=socks5',
    #     '--proxy-auth=YOUR_PASSWORD', 
    #     ]

    #Start up the webdriver with our service args. We're using PhantomJS since it is easy to add all this fake info into, and also can run headlessly while still loading JavaScript
    # driver = webdriver.PhantomJS(service_args=service_args)
    
    # Use the driver below if you don't want to use any of the fancy proxy stuff.
    driver = webdriver.Chrome()

    # Add gerneric cookies so we look even more legit
    # pickle.dump(driver.get_cookies(), open("cookies.pkl","wb"))

    return driver

def get_ip(driver):
    driver.get('http://www.whatsmyip.org/')
    soup = BS(driver.page_source, 'html.parser')
    ip = soup.find(id='ip').text
    print(ip)
    return ip

def write_results_local(data, data_type):
    #Write all results to a local file, to be mathced up against on future runs

    if data_type == 'links':
        f = open('loopnet_results.txt', 'w')
    elif data_type == 'addresses':
        f = open('loopnet_addresses.txt', 'w')
    simplejson.dump(data, f)
    f.close()

def read_past_results(filename=None):
    if filename is None:
        filename = 'loopnet_results.txt'
    try:
        f = open(filename, 'r')
        results = simplejson.load(f)
        f.close()
    except Exception as e:
        results = []
    return results

def add_entry(entry, data, data_type):
    if data_type == 'links':
        f = open('loopnet_results.txt', 'r')
    elif data_type == 'addresses':
        f = open('loopnet_addresses.txt', 'r')
    results = simplejson.load(f)
    f.close()
    results.append(entry)
    write_results_local(results, 'links')


def get_loopnet_results(driver):
    #Iterate through the results pages and grab all the specific URLs
    links = []
    for i in range(1,26):
        print(i)
        time.sleep(random.randint(5,15))
        # url = 'http://www.loopnet.com/for-sale/retail/{}/?sk=d29235b26a00d968bdf33458f6fa85dd&bb=69lskl-j0W_g3gyrq6yJ'.format(str(i))
        url = 'http://www.loopnet.com/for-sale/retail/{}/?bb=t8tpso19yU2xwwpm6vrJ'.format(str(i))
        driver.get(url)
        soup = BS(driver.page_source, 'html.parser')
        results = soup.find(class_='placards').find_all('article')
        for result in results:
            try:
                link = result.find(class_='listing-address').a['href']
                links.append(link)
            except AttributeError as e:
                print('Page data collected')
                break
    return links

def get_loopnet_result_data(driver, links):
    #Go through each loopnet result page and take all the information that we want and save it to be used in our Excel book later

    #Need to assign these all as blanks, since not all are always available and we don't want any undefined variables
    name = ''
    town = ''
    realator_first_name = ''
    realator_last_name = ''
    broker_company = ''
    phone_number = ''
    price = ''
    sq_ft = ''
    cap_rate = ''
    price_sf = ''
    property_type = ''
    property_subtype = ''
    tenancy = ''
    image_url = ''
    description = ''
    highlights = ''
    list_id = ''

    #create a list for our page data
    results_data = []
    #read the old results from last time, so we make sure we're only getting new info
    old_links = read_past_results()
    # old_addresses = read_past_results('addresses')
    bad_omen = False
    for p_count, link in enumerate(links):
        if bad_omen:
            break
        #if it's an old link, move along
        print(p_count)
        if link in old_links:
            print('Old link. Moving on')
            continue
        else:
            add_entry(link, old_links, 'links')
            print('Grabbing new property info')

        #Since we're loading this all from somewhere in Northern Europe (and their site has a lot of JavaScript), sometimes it takes 2 tries to load a page. Added a timeout as well. 
        try:
            with timeout(60*3, exception=RuntimeError):
                driver.get(link)
                print(link)
                time.sleep(random.randint(6,10))
                try:
                    soup = BS(driver.page_source, 'html.parser')
                except Exception as e:
                    time.sleep(10)
                    try:
                        driver.get(link)
                    except Exception as e:
                        output_to_excel(results_data)
                        bad_omen = True
                        break
                #Start off by grabbing the address/name and town. If we are unable to grab these, there is a 99% chance it is because their site has rate limited us (these fields are always available otherwise).
                try:
                    address_town = soup.find(class_='column-09 column-tiny-08').h1.text.split('\n')
                    address = address_town[0]
                    town = address_town[1].strip()[:-1].strip()
                except Exception as e:
                    print('Probably got rate limited...')
                    time.sleep(60)
                    continue
                try:
                    name = soup.find(class_='listing-name').text
                except Exception as e:
                    print('Name info not found: ' + str(e))
                    name = ''
                #These fields are static, so we'll always get this info
                try:
                    realator_first_name = soup.find(class_='name first-name').text
                    realator_last_name = soup.find(class_='name last-name').text
                except Exception as e:
                    print('Realator info not found: ' + str(e))
                    realator_first_name = ''
                    realator_last_name = ''
                try:
                    phone_number = soup.find(class_='center-wrap').p.text
                except Exception as e:
                    print('Phone info not found: ' + str(e))
                    phone_number = ''
                try:
                    broker_company = soup.find(class_='company-name').text
                except Exception as e:
                    print('Broker Company info not found: ' + str(e))
                    broker_company = ''

                #Results from this table dynamically change, and not all are always available, or in the same order
                results_table = soup.find(class_='property-data').tbody.find_all('tr')
                table_length = len(results_table)
                table_list = []
                #Here we grab everything from the table, and strip it down to just the text portions we want.
                for i in range(0, table_length):
                    td_list = results_table[i].find_all('td')
                    for item in td_list:
                        item = item.text.strip().replace('  ','').replace('\n', '')
                        if item != '':
                            table_list.append(item)

                #Here we take all our data from that table, and iterate through it to dynamically assign all the variables since that table is often different on each page
                count = 0
                for item in table_list:
                    count += 1
                    try:
                        if item == 'Price':
                            price = table_list[count] 
                        elif item == 'Building Size':
                            sq_ft = table_list[count] 
                        elif item == 'Cap Rate':
                            cap_rate = table_list[count]
                        elif item == 'Price/SF':
                            price_sf =  table_list[count]
                        elif item == 'Property Type':
                            property_type = table_list[count]
                        elif item == 'Property Sub-type':
                            property_subtype = table_list[count]
                        elif item == 'Tenancy':
                            tenancy = table_list[count]
                        else:
                            continue
                    except Exception as e:
                        print(e)
                        break
                try:
                    if float(cap_rate.replace('%','')) < 6.75 or property_type not in ['Retail', 'Industrial'] or tenancy != 'Single':
                        print('Parameters for collection not met. Moving on...')
                        continue
                except ValueError as e:
                    try:
                        cap_rate = non_decimal.sub('', cap_rate)
                        if float(cap_rate.replace('%','')) < 6.75 or property_type != 'Retail' or tenancy != 'Single':
                            print('Parameters for collection not met. Moving on...')
                            continue
                    except Exception as e:
                        print(e)
                        continue
                except Exception as e:
                    print(e)
                    continue
                #These next 3 items are seperate from the table, so their paths are static.
                try:
                    image_url = soup.find(class_='slide active ng-isolate-scope').img['src']
                except Exception as e:
                    print('Image URL info not found: ' + str(e))
                    image_url = ''
                try:    
                    description = soup.find(class_='description').find_all(class_='column-12')[1].text
                except Exception as e:
                    print('Description info not found: ' + str(e))
                    description = ''
                try:
                    highlights = soup.find(class_='highlights').find(class_='bulleted-list').text.strip().replace('\n', '. ')
                except Exception as e:
                    print('Highlight info not found: ' + str(e))
                    highlights = ''

                #grab the list id from the URL
                list_id = driver.current_url.split('/')[4]

                #make out dictionary, and add it to our list of results data.
                results_dict = {
                'name': name,
                'address': address,
                'town': town,
                'first_name': realator_first_name,
                'last_name': realator_last_name,
                'phone': phone_number,
                'price': price,
                'square_feet': sq_ft,
                'broker_company': broker_company,
                'cap_rate': cap_rate,
                'price_sq_ft': price_sf,
                'property_type': property_type,
                'property_subtype': property_subtype,
                'tenancy': tenancy,
                'image_url': image_url,
                'description': description,
                'highlights': highlights,
                'list_id': list_id,
                'url': driver.current_url
                }
                results_data.append(results_dict)
        except RuntimeError as e:
            print('Timeout')
            time.sleep(random.randint(15,30))
            continue

    if not bad_omen:
        return results_data
    else:
        results_data = False
        return results_data

def output_to_excel(results_data):
    #Here we compile the whole Excel book
    f = open('loopnet_excel_results.txt', 'w')
    simplejson.dump(results_data, f)
    f.close()
    #create the workbook
    wb = Workbook()
        #Add a new worksheet
    try:
            ws = wb.worksheets[0]
            ws.title = 'Results'
    except Exception as e:
        try:
            ws = wb.create_sheet('Results')
            ws = wb.worksheets[0]
        except Exception as e:
            print('Sheet creation error: ' + str(e))

    try:
        #Add the headers
        ws.cell(row=1, column=1).value = 'Listing Name'
        ws.cell(row=1, column=2).value = 'Address'
        ws.cell(row=1, column=3).value = 'Town'
        ws.cell(row=1, column=4).value = 'Price'
        ws.cell(row=1, column=5).value = 'Square Feet'
        ws.cell(row=1, column=6).value = 'Broker Company'
        ws.cell(row=1, column=7).value = 'First Name'
        ws.cell(row=1, column=8).value = 'Last Name'
        ws.cell(row=1, column=9).value = 'Phone'
        ws.cell(row=1, column=10).value = 'Cap Rate'
        ws.cell(row=1, column=11).value = 'Offer Cap'
        ws.cell(row=1, column=12).value = 'Calc Offer Cap'
        ws.cell(row=1, column=13).value = 'Calc Offer Price'
        ws.cell(row=1, column=14).value = 'Calc Rent'
        ws.cell(row=1, column=15).value = 'Price/Sqft'
        ws.cell(row=1, column=16).value = 'Property Type'
        ws.cell(row=1, column=17).value = 'Property Subtype'
        ws.cell(row=1, column=18).value = 'Tenancy'
        ws.cell(row=1, column=19).value = 'Description'
        ws.cell(row=1, column=20).value = 'Highlights'
        ws.cell(row=1, column=21).value = 'list ID'
        ws.cell(row=1, column=22).value = 'Image_URL'
        ws.cell(row=1, column=23).value = 'Page URL'
        print('Headers successfully written')
        
    except Exception as e:
        print('Cell info write error: ' + str(e))

    for i, data in enumerate(results_data):
        i+=2
        try:
            #write column 2
            ws.cell(row=i, column=1).value = data['name']
            ws.cell(row=i, column=2).value = data['address']
            ws.cell(row=i, column=3).value = data['town']
            ws.cell(row=i, column=4).value = data['price']
            ws.cell(row=i, column=5).value = data['square_feet']
            ws.cell(row=i, column=6).value = data['broker_company']
            ws.cell(row=i, column=7).value = data['first_name']
            ws.cell(row=i, column=8).value = data['last_name']
            ws.cell(row=i, column=9).value = data['phone']
            ws.cell(row=i, column=10).value = data['cap_rate']
            ws.cell(row=i, column=11).value = ''
            ws.cell(row=i, column=12).value = '=100*INT(K'+str(i)+'*10000)/10000'
            ws.cell(row=i, column=13).value = '=TEXT(INT((D'+str(i)+'*J'+str(i)+')/K'+str(i)+'),"#,###")'
            ws.cell(row=i, column=14).value = '=TEXT(INT(D'+str(i)+'*J'+str(i)+'),"#,###")'
            ws.cell(row=i, column=15).value = data['price_sq_ft']
            ws.cell(row=i, column=16).value = data['property_type']
            ws.cell(row=i, column=17).value = data['property_subtype']
            ws.cell(row=i, column=18).value = data['tenancy']
            ws.cell(row=i, column=19).value = data['description']
            ws.cell(row=i, column=20).value = data['highlights']
            ws.cell(row=i, column=21).value = data['list_id']
            ws.cell(row=i, column=22).value = data['image_url']
            ws.cell(row=i, column=23).value = data['url']
            print('Row '+str(i)+' written')
        except Exception as e:
            print('Cell info write error: ' + str(e))
            continue
    #Save and quit
    wb.save(filename='loopnet_results.xlsx')

#Run from main
if __name__ == '__main__':
    driver = start_driver()
    ip = get_ip(driver)
    links = get_loopnet_results(driver)
    # write_results_local(links) 
    # ^ Only uncomment this if this is the first time running this program on the machine
    results_data = get_loopnet_result_data(driver, links)
    print('Outputting to Excel')
    if results_data is not False:
        output_to_excel(results_data)
        print('Excel file saved.')
    else:
        print('Bad Omen detected. Breaking.')
    driver.quit()
    print('Complete')


















