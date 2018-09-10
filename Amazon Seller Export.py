from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup as BS
import time
import psycopg2
import re
from pushbullet import Pushbullet
import csv
import openpyxl
import codecs
import xlrd


def postgres_ref_db_connection():
    """
    Set up the postgres connection
    """
    return psycopg2.connect(host='HOST',
                          user='USERNAME',
                          password='PASSWORD',
                          database='DB_NAME',
                      port=5432)

def insert_records(insert_statement, values):
    """
    Inserts a list of values into a database

    param insert_statement: The SQL query used to insert a row of data
    type insert_statement: string representing a SQL query
    param values: the list of tuples to insert into the database
    type values: list containing a tuple
    """
    if len(values) > 0:
        db = postgres_ref_db_connection()

        cur = db.cursor()
        try:
            if len(values) >= 1:
                cur.executemany(insert_statement, values)
                db.commit()
                print("INSERTED INTO DB")
                    
        except Exception as err:
            print("EXCEPTION")
            print(err)
            db.rollback()
            pass
        finally:
            cur.close()
            db.close()

def create_driver(path_to_profile):
    profile = webdriver.FirefoxProfile(path_to_profile)
    driver = webdriver.Firefox(profile)
    return driver

def go_to_central_and_get_report(driver):
    driver.get('https://sellercentral.amazon.com/')
    time.sleep(4)
    try:
        driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div[1]/div/div[1]/a/kat-button/button').click()
        time.sleep(5)
    except Exception as e:
        pass
    try:
        driver.find_element_by_xpath('//*[@id="ap_email"]').send_keys('USERNAME')
        time.sleep(2)
    except Exception as e:
        pass
    try:
        driver.find_element_by_xpath('//*[@id="ap_password"]').send_keys('PASSWORD')
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="ap_password"]').submit()
        time.sleep(45)
        code = get_text()
        driver.find_element_by_xpath('//*[@id="auth-mfa-otpcode"]').send_keys(code)
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="auth-mfa-otpcode"]').submit()
        time.sleep(5)
    except Exception as e:
        pass
    driver.get('https://sellercentral.amazon.com/gp/ssof/reports/search.html#orderAscending=&recordType=AFNShipmentReport&noResultType=&merchantSku=&fnSku=&FnSkuXORMSku=&reimbursementId=&orderId=&genericOrderId=&asin=&lpn=&shipmentId=&problemType=ALL_DEFECT_TYPES&hazmatStatus=&inventoryEventTransactionType=&fulfillmentCenterId=&inventoryAdjustmentReasonGroup=&eventDateOption=1&fromDate=mm%2Fdd%2Fyyyy&toDate=mm%2Fdd%2Fyyyy&startDate=&endDate=&fromMonth=1&fromYear=2018&toMonth=1&toYear=2018&startMonth=&startYear=&endMonth=&endYear=')
    time.sleep(5)
    driver.find_element_by_xpath('/html/body/div[7]/div[3]/div[7]/div[2]/form/table/tbody/tr[1]/td[3]/select/option[2]').click()
    time.sleep(2)
    driver.find_element_by_xpath('/html/body/div[7]/div[3]/div[7]/div[2]/form/table/tbody/tr[4]/td[2]/button').click()
    for i in range(100):
        try:
            driver.find_element_by_xpath('/html/body/div[7]/div[3]/div[9]/table/tbody/tr[1]/td[4]/a/span').click()
            time.sleep(5)
            break
        except NoSuchElementException as e:
            time.sleep(10)
    soup = BS(driver.page_source, 'html.parser')
    filename = soup.find(class_='buttonImage')['href'].split('?')[0].split('__')[1]
    return filename

def get_text():
    pb = Pushbullet('PUSHBULLET API KEY')
    while True:
        pushes = pb.get_pushes()
        latest = pushes[0]['body']
        check = open('2fa.txt', 'r').read()
        if latest == check:
            time.sleep(10)
        else:
            open('2fa.txt', 'w').close()
            with open("2fa.txt", "w") as text_file:
                text_file.write(latest)
            break
    pb.delete_push(pushes[0].get("iden"))
    code = latest.replace(' is your Amazon security code.', '')
    return code

def convert_txt_to_xls(filename):
    input_file = filename
    output_file = 'seller_data.csv'

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    with open(input_file, 'rt', encoding='latin-1') as data:
        reader = csv.reader(data, delimiter='\t')
        for row in reader:
            try:
                ws.append(row)
            except Exception as e:
                continue

    wb.save(output_file)

def csv_to_db():
        try:
            book = xlrd.open_workbook("seller_data.csv")
            sheet = book.sheet_by_name("Sheet")

            query = """INSERT INTO SCHEMA.TABLE (Amazon_Order_Id, Shipment_ID, Shipment_Item_Id, Amazon_Order_Item_Id, Purchase_Date, Payments_Date, Shipment_Date, Reporting_Date, Buyer_Email, Buyer_Name, Merchant_SKU, Title, Shipped_Quantity, Currency, Item_Price, Item_Tax, Shipping_Price, Shipping_Tax, Gift_Wrap_Price, Gift_Wrap_Tax, Ship_Service_Level, Recipient_Name, Shipping_Address_1, Shipping_Address_2, Shipping_Address_3, Shipping_City, Shipping_State, Shipping_Postal_Code, Shipping_Country_Code, Item_Promo_Discount, Shipment_Promo_Discount, Carrier, Tracking_Number, Estimated_Arrival_Date, FC, Fulfillment_Channel, Sales_Channel) VALUES (%s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

            # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
            values_list = []
            for r in range(1, sheet.nrows):
                    Amazon_Order_Id = sheet.cell(r,0).value
                    if Amazon_Order_Id == 'Amazon Order Id':
                        continue
                    Shipment_ID = sheet.cell(r,2).value
                    Shipment_Item_Id = sheet.cell(r,3).value
                    Amazon_Order_Item_Id = sheet.cell(r,4).value
                    Purchase_Date = sheet.cell(r,6).value
                    Payments_Date = sheet.cell(r,7).value
                    Shipment_Date = sheet.cell(r,8).value
                    Reporting_Date = sheet.cell(r,9).value
                    Buyer_Email = sheet.cell(r,10).value
                    Buyer_Name = sheet.cell(r,11).value
                    Merchant_SKU = sheet.cell(r,13).value
                    Title = sheet.cell(r,14).value
                    Shipped_Quantity = sheet.cell(r,15).value
                    Currency = sheet.cell(r,16).value
                    Item_Price = sheet.cell(r,17).value
                    Item_Tax = sheet.cell(r,18).value
                    Shipping_Price = sheet.cell(r,19).value
                    Shipping_Tax = sheet.cell(r,20).value
                    Gift_Wrap_Price = sheet.cell(r,21).value
                    Gift_Wrap_Tax = sheet.cell(r,22).value
                    Ship_Service_Level = sheet.cell(r,23).value
                    Recipient_Name = sheet.cell(r,24).value
                    Shipping_Address_1 = sheet.cell(r,25).value
                    Shipping_Address_2 = sheet.cell(r,26).value
                    Shipping_Address_3 = sheet.cell(r,27).value
                    Shipping_City = sheet.cell(r,28).value
                    Shipping_State = sheet.cell(r,29).value
                    Shipping_Postal_Code = sheet.cell(r,30).value
                    Shipping_Country_Code = sheet.cell(r,31).value
                    Item_Promo_Discount = sheet.cell(r,40).value
                    Shipment_Promo_Discount = sheet.cell(r,41).value
                    Carrier = sheet.cell(r,42).value
                    Tracking_Number = sheet.cell(r,43).value
                    Estimated_Arrival_Date = sheet.cell(r,44).value
                    FC = sheet.cell(r,45).value
                    Fulfillment_Channel = sheet.cell(r,46).value
                    Sales_Channel = sheet.cell(r,47).value

                    # Assign values from each row
                    values = (Amazon_Order_Id, Shipment_ID, Shipment_Item_Id, Amazon_Order_Item_Id, Purchase_Date, Payments_Date, Shipment_Date, Reporting_Date, Buyer_Email, Buyer_Name, Merchant_SKU, Title, Shipped_Quantity, Currency, Item_Price, Item_Tax, Shipping_Price, Shipping_Tax, Gift_Wrap_Price, Gift_Wrap_Tax, Ship_Service_Level, Recipient_Name, Shipping_Address_1, Shipping_Address_2, Shipping_Address_3, Shipping_City, Shipping_State, Shipping_Postal_Code, Shipping_Country_Code, Item_Promo_Discount, Shipment_Promo_Discount, Carrier, Tracking_Number, Estimated_Arrival_Date, FC, Fulfillment_Channel, Sales_Channel)
                    values_list.append(values)
            insert_records(query, values_list)
            values_list = []
        except Exception as e:
            print(e) 

if __name__ == "__main__":
    driver = create_driver('ff_profile')
    filename = go_to_central_and_get_report(driver)
    driver.quit()
    convert_txt_to_xls(filename)
    csv_to_db()